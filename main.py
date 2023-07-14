from datetime import date, datetime
from openpyxl import Workbook, load_workbook
import qrcode
import cv2
import pyzbar.pyzbar as pyzbar

# Function to generate QR code for a student
def generate_qr_code(student_id, student_name):
    data = f"ID: {student_id}\nName: {student_name}"
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(data)
    qr.make(fit=True)
    qr_image = qr.make_image(fill="black", back_color="white")
    qr_image.save(f"qr_codes/{student_id}.png")

# Function to scan QR code and record attendance
def scan_qr_code():
    cap = cv2.VideoCapture(0)
    font = cv2.FONT_HERSHEY_PLAIN

    while True:
        _, frame = cap.read()
        decoded_objects = pyzbar.decode(frame)

        for obj in decoded_objects:
            data = obj.data.decode('utf-8')
            process_attendance(data)

        cv2.imshow("QR Code Scanner", frame)
        key = cv2.waitKey(1)
        if key == 27:
            break

    cap.release()
    cv2.destroyAllWindows()

# Function to process attendance and save it in Excel
def process_attendance(data):
    student_id = ""
    student_name = ""
    attendance_date = date.today().strftime("%Y-%m-%d")
    attendance_status = "Present"

    # Parse the QR code data to extract student ID and name
    for line in data.split('\n'):
        if line.startswith("ID:"):
            student_id = line.split("ID:")[1]
        elif line.startswith("Name:"):
            student_name = line.split("Name:")[1]

    # Load the attendance Excel file
    workbook = load_workbook("attendance.xlsx")
    sheet = workbook.active

    # Check if the student ID is already recorded for today's date
    for row in sheet.iter_rows(values_only=True):
        if row[0] == student_id and row[2] == attendance_date:
            # Get the existing timestamps and append the current timestamp
            timestamps = row[3]
            timestamps += ", " + datetime.now().strftime("%H:%M:%S")
            sheet.cell(row=row[4], column=4).value = timestamps
            workbook.save("attendance.xlsx")
            print("Attendance already recorded. Additional timestamp added.")
            return

    # Add a new row for attendance
    new_row = [student_id, student_name, attendance_date, datetime.now().strftime("%H:%M:%S"), sheet.max_row]
    sheet.append(new_row)

    # Save the attendance Excel file
    workbook.save("attendance.xlsx")
    print("Attendance recorded successfully.")

# Main program
if __name__ == "__main__":
    # Generate QR codes for students
    students = [
        {"id": "001", "name": "John Doe"},
        {"id": "002", "name": "Jane Smith"},
        {"id": "003", "name": "Harsh"}
        # Add more students as needed
    ]

    for student in students:
        generate_qr_code(student["id"], student["name"])

    # Start the QR code scanning process
    scan_qr_code()
